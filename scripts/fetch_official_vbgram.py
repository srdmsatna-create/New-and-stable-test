#!/usr/bin/env python3
"""Fetch VB-G RAM G official MIS data using a real browser.

Goal: obtain the same Daily Report workbook (RepDay + Sheet1 + VBG) without
manual upload. The script is defensive: it never publishes partial/bad data.

Optional GitHub secrets/env:
  VBGRAM_DAILY_REPORT_URL : direct report page URL, if known
  VBGRAM_USERNAME / VBGRAM_PASSWORD : only if official site prompts login
  VBGRAM_COOKIE : raw Cookie header if a session cookie is required

The committed defaults are the official links already shown in the portal.
"""
import os, re, json, shutil, sys, time
from pathlib import Path
from datetime import datetime, timezone

ROOT=Path(__file__).resolve().parents[1]
INCOMING=ROOT/'incoming'; RAW=ROOT/'raw'; DATA=ROOT/'data'
INCOMING.mkdir(exist_ok=True); RAW.mkdir(exist_ok=True); DATA.mkdir(exist_ok=True)

HOME='https://vbgramg.dord.gov.in/vbgramg/home.aspx'
MIS='https://vbgramgrep.dord.gov.in/VBGRAMG/MISreport.aspx'
ONGOING='https://vbgramgrep.dord.gov.in/VBGRAMG/dynamic_work_details.aspx?payload=4PmH2eRA9khYNUNqz1h5yt9D8POKLA7Afp0nercX3xt22K65u-hNco55SZiMHr78IufQr-Pyxw1-2tJEz-65UMtG5kOTBzCEHurJmRrAtoAIfVSTK-qhJdX02vLZMWrVbwM-oS9xX58g6SiO5ODhhFid9RqKvnwTnS-hLkXfa1-25phIp66JlphIcilUU7cK'
REPORT_URL=os.environ.get('VBGRAM_DAILY_REPORT_URL','').strip() or MIS
DIRECT_XLSX_URL=os.environ.get('VBGRAM_DAILY_REPORT_XLSX_URL','').strip() or os.environ.get('DAILY_REPORT_XLSX_URL','').strip()
DIRECT_ONGOING_CSV_URL=os.environ.get('VBGRAM_ONGOING_CSV_URL','').strip()
USERNAME=os.environ.get('VBGRAM_USERNAME','').strip()
PASSWORD=os.environ.get('VBGRAM_PASSWORD','').strip()
COOKIE=os.environ.get('VBGRAM_COOKIE','').strip()

status={'startedAt':datetime.now(timezone.utc).isoformat(),'ok':False,'steps':[],'source':'Official VB-G RAM G'}
def note(step,ok=True,detail=''):
    status['steps'].append({'step':step,'ok':bool(ok),'detail':str(detail)[:1000]})
    print(('OK ' if ok else 'WARN ')+step+(' :: '+str(detail) if detail else ''))

def save_status():
    status['finishedAt']=datetime.now(timezone.utc).isoformat()
    (DATA/'fetch-status.json').write_text(json.dumps(status,ensure_ascii=False,indent=2),encoding='utf-8')
    (ROOT/'auto-status.js').write_text('window.AUTO_FETCH_STATUS='+json.dumps(status,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')


def direct_download(url, name):
    if not url: return None
    try:
        import urllib.request
        req=urllib.request.Request(url,headers={'User-Agent':'Mozilla/5.0','Cache-Control':'no-cache','Pragma':'no-cache'})
        with urllib.request.urlopen(req,timeout=90) as r:
            b=r.read()
        if len(b)<500:
            raise ValueError(f'download too small: {len(b)} bytes')
        target=RAW/name
        target.write_bytes(b)
        note(f'direct source downloaded: {name}',True,f'{len(b)} bytes')
        return target
    except Exception as e:
        note(f'direct source download failed: {name}',False,e)
        return None

# V51 primary path: direct machine-readable sources, when configured.
# This avoids CAPTCHA/browser fragility. Browser scraping remains a fallback.
direct_workbook=direct_download(DIRECT_XLSX_URL,'Daily_Report_direct.xlsx')
direct_ongoing=direct_download(DIRECT_ONGOING_CSV_URL,'Ongoing_Works_direct.csv')
if not DIRECT_XLSX_URL:
    note('direct Daily Report source configured',False,'GitHub secret VBGRAM_DAILY_REPORT_XLSX_URL is empty; unattended fresh workbook update cannot be guaranteed when MIS requires CAPTCHA/session.')
if not DIRECT_ONGOING_CSV_URL:
    note('direct Ongoing CSV source configured',False,'GitHub secret VBGRAM_ONGOING_CSV_URL is empty; work-level CSV may remain on the previous committed snapshot.')

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except Exception as e:
    note('playwright import',False,e); save_status(); raise

def maybe_login(page):
    # Generic login support only when credentials are supplied.
    if not (USERNAME and PASSWORD): return False
    pw=page.locator('input[type="password"]')
    if pw.count()==0: return False
    user=page.locator('input[type="text"],input[type="email"]')
    if user.count(): user.first.fill(USERNAME)
    pw.first.fill(PASSWORD)
    btn=page.get_by_role('button',name=re.compile('login|sign in|submit|प्रवेश',re.I))
    if btn.count(): btn.first.click()
    else: pw.first.press('Enter')
    page.wait_for_timeout(2000)
    note('login submitted',True)
    return True

def table_dump(page,label):
    try:
        tables=page.locator('table')
        out=[]
        for i in range(min(tables.count(),60)):
            t=tables.nth(i)
            rows=[]
            trs=t.locator('tr')
            for r in range(min(trs.count(),3000)):
                cells=trs.nth(r).locator('th,td')
                rows.append([cells.nth(c).inner_text().strip() for c in range(cells.count())])
            if rows: out.append(rows)
        (RAW/f'{label}_tables.json').write_text(json.dumps(out,ensure_ascii=False),encoding='utf-8')
        note(f'{label} tables captured',True,len(out))
    except Exception as e: note(f'{label} tables captured',False,e)

def try_download(page):
    """Try common Excel/export controls; return downloaded path or None."""
    # buttons/links with likely export text
    pats=re.compile(r'excel|xlsx|download|export|डाउनलोड|एक्सेल',re.I)
    candidates=page.get_by_role('link',name=pats)
    if candidates.count()==0: candidates=page.get_by_role('button',name=pats)
    for i in range(min(candidates.count(),25)):
        el=candidates.nth(i)
        try:
            if not el.is_visible(): continue
            with page.expect_download(timeout=12000) as di:
                el.click()
            d=di.value
            suggested=d.suggested_filename or 'Daily Report.xlsx'
            if not suggested.lower().endswith(('.xlsx','.xls','.csv')): suggested+=' .xlsx'
            target=RAW/suggested.replace('/','_')
            d.save_as(str(target))
            note('official export downloaded',True,target.name)
            return target
        except Exception:
            continue
    return None

def validate_workbook(path):
    if not path or path.suffix.lower() not in ('.xlsx','.xlsm','.xls'): return False
    try:
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True)
        need={'RepDay','Sheet1','VBG'}
        ok=need.issubset(set(wb.sheetnames))
        note('workbook sheet validation',ok,','.join(wb.sheetnames[:20]))
        wb.close(); return ok
    except Exception as e:
        note('workbook sheet validation',False,e); return False



def parse_official_summary_from_tables():
    """V46 fallback: Screen-2 is authoritative for shared daily KPIs.
    A wider rich table may supply ongoing/category columns. If that wider table is
    unavailable, preserve the previous valid rich values instead of substituting
    incompatible Screen-2 columns.
    """
    source=RAW/'mis_tables.json'
    if not source.exists(): return None
    try: tables=json.loads(source.read_text(encoding='utf-8'))
    except Exception as e: note('official table parse',False,e); return None
    valid={'AMARPATAN','MAIHAR','RAMNAGAR','MAJHGAWAN','NAGOD','RAMPUR BAGHELAN','SATNA','UNCHAHARA'}
    order=['AMARPATAN','MAIHAR','RAMNAGAR','MAJHGAWAN','NAGOD','RAMPUR BAGHELAN','SATNA','UNCHAHARA']
    rich_fields=['totalGP','musterGP','dysfunctionalGP','labourAll','mrAll','ongoingAll','labourIndividual','mrIndividual','labourCommunity','mrCommunity','pmayOngoing','pmayMR','ekLabour','ekOngoing','ekMR']
    rich_idx=[2,3,4,6,7,8,10,11,12,13,15,16,18,19,20]
    def n(v):
        try:return float(str(v).replace(',','').strip() or 0)
        except:return 0.0
    # Load previous rich values as safe fallback.
    previous={}
    old=DATA/'official-summary.csv'
    if old.exists():
        try:
            import csv
            with old.open(encoding='utf-8-sig',newline='') as f:
                previous={str(r.get('janpad','')).strip().upper():dict(r) for r in csv.DictReader(f)}
        except Exception: previous={}
    best_rich={}; best_screen={}
    for table in tables:
        rich={}; screen={}
        for r in table:
            if len(r)<2: continue
            jan=str(r[1]).strip().upper().replace('RAMPUR\nBAGHELAN','RAMPUR BAGHELAN')
            jan=' '.join(jan.split())
            if jan not in valid: continue
            if len(r)>=21:
                z={'janpad':jan}
                for k,i in zip(rich_fields,rich_idx): z[k]=n(r[i])
                if z['totalGP']>0 and z['ongoingAll']>0: rich[jan]=z
            # Screen-2 row: SNo, Block, Total GP, GP with progress, Labour, Works MR, no-eKYC, Muster Rolls
            if len(r)>=8:
                z={'janpad':jan,'totalGP':n(r[2]),'musterGP':n(r[3]),'labourAll':n(r[4]),'mrAll':n(r[5]),'noEkyc':n(r[6]),'mrs':n(r[7])}
                if z['totalGP']>0 and z['musterGP']>=0 and z['mrs']>=0: screen[jan]=z
        if len(rich)>len(best_rich): best_rich=rich
        if len(screen)>len(best_screen): best_screen=screen
    if set(best_screen)!=valid:
        note('Screen-2 table parse',False,f'Expected 8 Janpads, got {len(best_screen)}'); return None
    fields=['janpad','totalGP','musterGP','dysfunctionalGP','labourAll','mrAll','noEkyc','mrs','ongoingAll','labourIndividual','mrIndividual','labourCommunity','mrCommunity','pmayOngoing','pmayMR','ekLabour','ekOngoing','ekMR']
    result=[]
    for j in order:
        base={k:n(previous.get(j,{}).get(k,0)) for k in fields if k!='janpad'}
        if j in best_rich: base.update(best_rich[j])
        sc=best_screen[j]
        # Screen-2 wins for every shared metric.
        base.update(sc); base['janpad']=j; base['dysfunctionalGP']=max(0,base['totalGP']-base['musterGP'])
        result.append(base)
    import csv
    out=DATA/'official-summary.csv'
    with out.open('w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(result)
    note('Screen-2 table parse',True,'8 Janpads authoritative; rich columns preserved/merged')
    return out

with sync_playwright() as p:
    browser=p.chromium.launch(headless=True,args=['--no-sandbox'])
    context=browser.new_context(accept_downloads=True,viewport={'width':1600,'height':1000},extra_http_headers={'Cache-Control':'no-cache','Pragma':'no-cache'})
    if COOKIE:
        # Parse simple Cookie header into cookies for both official hosts.
        cookies=[]
        for pair in COOKIE.split(';'):
            if '=' not in pair: continue
            k,v=pair.strip().split('=',1)
            for domain in ['vbgramgrep.dord.gov.in','vbgramg.dord.gov.in']:
                cookies.append({'name':k,'value':v,'domain':domain,'path':'/'})
        try: context.add_cookies(cookies); note('session cookie loaded',True,len(cookies))
        except Exception as e: note('session cookie loaded',False,e)
    page=context.new_page()

    # Home warms cookies/session.
    try:
        r=page.goto(HOME,wait_until='domcontentloaded',timeout=60000)
        note('Home opened',bool(r and r.ok),getattr(r,'status',None))
    except Exception as e: note('Home opened',False,e)

    # Capture MIS and Ongoing independently. V50 deliberately keeps two exports:
    # one workbook for RepDay/Sheet1/VBG, and one dynamic_work_details CSV for
    # work-level Ongoing details. Previously the first download blocked the second.
    downloaded=direct_workbook if validate_workbook(direct_workbook) else None
    ongoing_download=direct_ongoing
    for url,label in [(REPORT_URL,'mis'),(ONGOING,'ongoing')]:
        try:
            r=page.goto(url,wait_until='domcontentloaded',timeout=90000)
            note(f'{label} opened',bool(r and r.ok),getattr(r,'status',None))
            maybe_login(page)
            page.wait_for_timeout(2500)
            table_dump(page,label)
            html=page.content(); (RAW/f'{label}.html').write_text(html,encoding='utf-8')
            got=try_download(page)
            if label=='mis' and not downloaded: downloaded=got
            elif label=='ongoing' and not ongoing_download: ongoing_download=got
        except Exception as e:
            note(f'{label} opened',False,e)

    # Search any captured browser responses/downloads is not needed when export worked.
    browser.close()

# V50: install the official dynamic_work_details work list independently of workbook mode.
# Prefer the official CSV/XLSX export; if the export control is unavailable, rebuild
# a CSV from the captured HTML table. This removes the old dated packaged snapshot.
def install_ongoing_export(path):
    required=['Janpad / Block Name','Panchayat Name','Work Code','Work Name','Work Status']
    dest=DATA/'Ongoing_Works_dynamic_work_details_latest.csv'
    try:
        rows=None
        if path and path.suffix.lower()=='.csv':
            import csv
            with path.open('r',encoding='utf-8-sig',newline='') as f:
                rd=csv.DictReader(f)
                if not set(required).issubset(set(rd.fieldnames or [])):
                    raise ValueError('expected dynamic_work_details headers not found in CSV')
                rows=list(rd); fields=rd.fieldnames
        elif path and path.suffix.lower() in ('.xlsx','.xlsm'):
            from openpyxl import load_workbook
            wb=load_workbook(path,read_only=True,data_only=True)
            found=None
            for ws in wb.worksheets:
                vals=list(ws.iter_rows(values_only=True))
                if not vals: continue
                hdr=[str(x).strip() if x is not None else '' for x in vals[0]]
                if set(required).issubset(set(hdr)):
                    found=(hdr,vals[1:]); break
            wb.close()
            if not found: raise ValueError('expected dynamic_work_details headers not found in workbook')
            fields,raw=found
            rows=[dict(zip(fields,r)) for r in raw if any(x not in (None,'') for x in r)]
        else:
            # HTML table fallback from Playwright capture.
            source=RAW/'ongoing_tables.json'
            if source.exists():
                tables=json.loads(source.read_text(encoding='utf-8'))
                for t in tables:
                    if not t: continue
                    hdr=[str(x).strip() for x in t[0]]
                    if set(required).issubset(set(hdr)):
                        fields=hdr
                        rows=[dict(zip(fields,r+['']*(len(fields)-len(r)))) for r in t[1:] if any(str(x).strip() for x in r)]
                        break
        if not rows:
            note('ongoing work export validation',False,'no usable CSV/XLSX/table found'); return None
        import csv
        with dest.open('w',encoding='utf-8-sig',newline='') as f:
            w=csv.DictWriter(f,fieldnames=fields,extrasaction='ignore'); w.writeheader(); w.writerows(rows)
        # lightweight row validation
        with dest.open('r',encoding='utf-8-sig',newline='') as f:
            rd=csv.DictReader(f); count=sum(1 for _ in rd)
        if count<=0:
            dest.unlink(missing_ok=True); note('ongoing work export validation',False,'zero data rows'); return None
        status['ongoingCsv']=str(dest.relative_to(ROOT))
        status['ongoingRows']=count
        note('fresh ongoing work CSV installed',True,f'{dest} rows={count}')
        return dest
    except Exception as e:
        note('ongoing work export validation',False,e); return None

ongoing_csv=install_ongoing_export(ongoing_download)

# V49: LIVE PORTAL FIRST. Always parse the current HTML Screen-2 table.
# A downloaded workbook is still useful for RepDay/VBG/rich fields, but it must NEVER
# overwrite fresher Screen-2 values shown on the official portal.
summary=parse_official_summary_from_tables()
workbook_ok=bool(downloaded and validate_workbook(downloaded))
if workbook_ok:
    dest=INCOMING/'Daily Report.xlsx'
    shutil.copy2(downloaded,dest)
    status['workbook']=str(dest.relative_to(ROOT))
    note('fresh Daily Report installed',True,dest)

if workbook_ok and summary:
    status['ok']=True; status['updateMode']='workbook+summary'; status['summary']=str(summary.relative_to(ROOT))
    note('live portal authority enabled',True,'Workbook supplies drill-down/rich fields; current portal Screen-2 overrides all shared KPIs')
elif summary:
    status['ok']=True; status['updateMode']='summary'; status['summary']=str(summary.relative_to(ROOT))
    note('live summary mode enabled',True,'Current official Screen-2 refreshes cards/table; previous rich fields are preserved only where portal does not expose them')
elif workbook_ok:
    status['ok']=True; status['updateMode']='workbook'
    note('workbook-only mode',True,'Live Screen-2 table was unavailable; workbook used as fallback')
else:
    status['ok']=False
    note('fresh official data installed',False,'Neither a valid workbook nor the live 8-Janpad official Screen-2 table was detected. Previous valid report remains live.')

save_status()
if not status['ok']:
    sys.exit(2)
