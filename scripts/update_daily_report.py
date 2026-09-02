#!/usr/bin/env python3
import os, json, re, urllib.request, tempfile
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'auto-data.js'
INCOMING=ROOT/'incoming'/'Daily Report.xlsx'
ROOT_XLSX=ROOT/'Daily Report.xlsx'
URL=os.environ.get('DAILY_REPORT_XLSX_URL','').strip()
VALID_JANPADS={'AMARPATAN','MAIHAR','RAMNAGAR','MAJHGAWAN','NAGOD','RAMPUR BAGHELAN','SATNA','UNCHAHARA'}

def n(v):
    try:return float(v or 0)
    except:return 0.0

def c(v): return '' if v is None else str(v).strip()
def up(v): return c(v).upper()
def janpad(v):
    j=up(v)
    return 'SATNA' if j=='SOHAWAL' else j



def final_category(name, wt, fy):
    name=c(name); wt=c(wt); t=(name+' '+wt).lower(); start=name.lower().strip()
    def has(p): return re.search(p,t,re.I) is not None
    if re.search(r'\bpmay\b|pmay[- ]?g|pradhan\s*mantri\s*awas',wt,re.I) or re.search(r'\bpmay\b|pmay[- ]?g',name,re.I):
        return 'PMAY-G'
    inst=re.compile(r'(school|prathmik|madhyamik|shala|vidyalaya|प्राथमिक|माध्यमिक|शाला|विद्यालय)',re.I)
    ek=re.compile(r'^(?:ek\s+(?:ma+a?|maa|mother)?\s*ki?\s*)?(?:bagiya|bagia)\b|^ek\s+bagiya\b|^ek\s+(?:maa?|ma|mाँ|मॉ|माॅ)\s+ke\s+naam\b|^(?:एक\s+)?(?:माँ|मा|मॉ|माॅ)\s*(?:की|के\s+नाम)?\s*बगिया\b|^एक\s+बगिया\b',re.I)
    if c(fy) in ('2025-2026','2026-2027') and ek.search(name) and not inst.search(name): return 'Ek Bagiya'
    rules=[
      ('Boundary Wall',r'(?:boundary|boundry|boundri|baundri|baundry|bawandri|bawndri|baudri|baundary|baun?dr[iy]|बाउंड्री|बाउण्डी|बाउण्?डी|बाउणडी|बाउंडरी|बॉउंड्री)'),
      ('Pulya',r'(?:puliya|pulia|pulya|पुलिया|culvert|cluvert|hume\s*pipe|क्रांस|\bcross\b)'),
      ('Cement Concrete',r'(?:\bpcc\b|\bcc\b|cement\s*concrete|concrete|nali|नाली|drain|drainage|grey\s*water|greywater|ग्रेवाटर|cover(?:e)?d\s*nali|rural\s*connectivity)'),
      ('Gravel Road',r'(?:gravel|greval|grewal|graval|grable|ग्रेवल|ग्रेबल|गे्रवल|गे्वल|mitti\s*mur+am|mur+am|murram|मिट्टी\s*मुरुम|मुरुम|sudoor|sudur|सुदूर\s*सड़क|\bbt\s*road\b|बीटी\s*रोड|bitumen)'),
      ('Water conservation & recharge',r'(?:\br\.?m\.?s\.?\b|आर\.?एम\.?एस|stop\s*dam|stap\s*dam|स्टाप\s*डैम|स्टापडैम|स्टाप\s*डेम|check\s*dam|चेक\s*डैम|ring\s*bund|d\s*frame|d\s*band|percolation|parkulation|पार्कुलेशन|hand\s*pump\s*recharge|handpump\s*recharge|hundpump\s*recharge)'),
      ('Farm Pond',r'(?:khet\s*talab|खेत\s*तालाब|\bctr\b.*khet\s*talab|farm\s*pond)'),
      ('Watershed Related Works',r'(?:dug\s*pond|डुग\s*पोंड|डग\s*पोंड|soak\s*pit|soakpit|shokpit|शोकपिट|gabion|\brfr\b|nadi\s*restoration|gully\s*plug|medh\s*bandhan|loose\s*boulder|new\s*talab)'),
      ('Dug Well Recharge',r'(?:samuday(?:i|ik|ak)\s*koop|community\s*well|सामुदायिक\s*कूप|dug\s*well\s*recharge|dugwell\s*recharge)'),
      ('Gap Filling in Plantation',r'(?:charagah|charagaah|chara\s*gah|chara\s*gaah|चारागाह|चारगाह|चारा\s*गाह|posh?an\s*vatika|vasudha|vashudha|bsudha|vassudha|वसुधा|vriksharopan|vraksha\s*ropan|bracharopan|braksharopan|वृक्षारोपण|व़क्षारोपण|वुक्षारोपण|faloudd?yan|faloudyan|फलोउद्यान|फलोउद्ययान|plantation|land\s*development)'),
      ('Crematorium',r'(?:shanti\s*dham|santi\s*dham|santhi\s*dham|shathi\s*dham|shanti\s*daham|shantidham|shantidam|mukti\s*dham|muktidham|मुक्तिधाम|शांति\s*धाम|शांती\s*धाम|शान्ति\s*धाम|शान्तिधाम)'),
      ('Panchayat and Community Hall',r'(?:panchayat\s*bhavan|panchyat\s*bhawan|panchyat\s*bhavan|naveen\s*panchayat|mangal\s*bhawan|samudai?k\s*bhawan|samudayik\s*bhavan|samudaik\s*bhavan|sabhagar\s*nirman|kaushal|कौशल|मंगल\s*भवन|पंचायत\s*भवन|सामुदायिक\s*भवन)'),
      ('SBM Works',r'(?:samudai?k\s*shauchalay|samudayik\s*shauchalay|सामुदायिक\s*शौचालय|segregation|segragation|kachara|karchra|kooda|kuda|कचरा|कूड़ा|\bnadep\b)'),
      ('Play Field',r'(?:play\s*ground|playground|khel\s*maidan|खेल\s*मैदान|खेल\s*का\s*मैदान|खेलो\s*के\s*मैदान)'),
    ]
    for cat,pat in rules:
        if has(pat): return cat
    ang=re.compile(r'^(?:a+nganwadi|a+ganwadi|a+anganbadi|a+ganbadi|anganbadi|aganbadi|आंगनवाडी|आंगनवाड़ी|आंगनवाड़ी|आंगनबाडी|आंगनबाड़ी|आंगनबाड़ी|आगनबाडी|आगनबाड़ी|आगनबाड़ी|आगनवाडी|आगनवाड़ी|आगनवाड़ी|आगवाड़ी|आगवाडी|ऑगनबाडी|ऑगनबाड़ी|ऑगनबाड़ी|ऑगनवाडी|ऑगनवाड़ी|ऑगनवाड़ी|अांगनवाडी|आंंगनवाडी)\b',re.I)
    if ang.search(start) and has(r'(?:bhavan|bhawan|nirman|kendra|sewa\s*kendra|शेष\s*कार्य|कार्य|भवन|निर्माण|केंद्र)'): return 'Anganwadi'
    if has(r'(?:kapil\s*dhara|kapildhara|कपिलधारा|open\s*dug\s*well|open\s*dugwell)'): return 'Kapildhara'
    if has(r'(?:cattle\s*shed|goat\s*shelter|poultry|pasu\s*shed|pashu\s*shed|पशु\s*शेड)'): return 'Poultry Cattle and Goat Shelter'
    if has(r'(?:micro\s*irrigation|सूक्ष्म\s*सिंचाई)'): return 'Irrigation infrastructure'
    if has(r'(?:uchit\s*mul|उचित\s*मूल|\bpds\b|food\s*grain|foodgrain|khad[hy]*yan|खाद्यान|खय्दायन|godam|gaushala|gau\s*shala|chabut|चबुत|paver|pever|pewar|retaining\s*wall|retarning\s*wall|ghat\s*nirman|bus\s*stop|pani\s*tanki|water\s*tank|kitchen\s*shed|laboratory|प्रयोगशाला|sub\s*health|उप\s*स्वास्थ्य|rangmanch|park\b|paper)'): return 'Other Works'
    rw=wt.lower()
    fallback=[
      ('Crematorium',r'crematorium'),('Gap Filling in Plantation',r'plantation|chara\s*gaah'),('Water conservation & recharge',r'check\s*dam|percolation'),
      ('Farm Pond',r'khet\s*talab|farm\s*pond'),('Watershed Related Works',r'dug\s*pond|soak\s*pit|gully\s*plug'),('Play Field',r'play\s*ground'),
      ('Cement Concrete',r'pcc|coverd\s*nali'),('Pulya',r'puliya|pulya'),('Boundary Wall',r'boundary'),('Poultry Cattle and Goat Shelter',r'cattle\s*shed'),('SBM Works',r'nadep|segregation|\bcsc\b'),('Water conservation & recharge',r'old\s*water|amrit\s*sarovar|roof\s*top'),('Watershed Related Works',r'contour\s*trench|loose\s*bolder|loose\s*boulder|nala\s*trench'),('Panchayat and Community Hall',r'panchayat.*community.*bhawan|community.*bhawan'),('Other Works',r'other')]
    for cat,pat in fallback:
        if re.search(pat,rw,re.I): return cat
    return wt or 'Other Works'
def obtain():
    if URL:
        fd,name=tempfile.mkstemp(suffix='.xlsx'); os.close(fd)
        req=urllib.request.Request(URL,headers={'User-Agent':'Mozilla/5.0 DailyReportBot/1.1'})
        with urllib.request.urlopen(req,timeout=90) as r, open(name,'wb') as f:f.write(r.read())
        return Path(name),URL
    if INCOMING.exists(): return INCOMING,'repo:incoming/Daily Report.xlsx'
    if ROOT_XLSX.exists(): return ROOT_XLSX,'repo:Daily Report.xlsx'
    raise SystemExit('No source workbook. Set DAILY_REPORT_XLSX_URL secret or add incoming/Daily Report.xlsx')

def extract_date(text):
    m=re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',c(text))
    return f'{int(m.group(1)):02d}-{int(m.group(2)):02d}-{m.group(3)}' if m else None

def parse(path):
    wb=load_workbook(path,data_only=True,read_only=True)
    for sh in ('RepDay','Sheet1','VBG'):
        if sh not in wb.sheetnames: raise SystemExit(f'Missing sheet: {sh}')

    # RepDay = Engineer / Cluster / GP drill-down only.
    rep=wb['RepDay']; repvals=list(rep.iter_rows(values_only=True))
    title=c(repvals[0][0]) if repvals else ''
    rep_date=extract_date(title); rows=[]; gpmap={}
    for r in repvals[3:]:
        if len(r)<12 or not c(r[0]) or not c(r[1]) or not c(r[5]): continue
        z={'janpad':janpad(r[0]),'engineer':c(r[1]),'cluster':c(r[2]),'ongoing':n(r[3]),'panchayat':up(r[5]),'gps':n(r[6]),'gpsProgress':n(r[7]),'labour':n(r[8]),'worksMR':n(r[9]),'noEkyc':n(r[10]),'mrs':n(r[11])}
        rows.append(z); gpmap[(z['janpad'],z['panchayat'])]=(z['engineer'],z['cluster'])
    if not rows: raise SystemExit('RepDay has no usable GP rows')

    # Sheet1 upper table = official work-load/category metrics (including total ongoing works).
    s1=wb['Sheet1']; svals=list(s1.iter_rows(values_only=True))
    sheet1_title=c(svals[0][0]) if svals else ''
    sheet1_date=extract_date(sheet1_title); official=[]
    for idx in (3,4,5,7,8,9,10,11):  # Excel rows 4,5,6,8..12
        if idx>=len(svals): continue
        r=svals[idx]
        if len(r)<20 or not c(r[1]): continue
        official.append({'janpad':janpad(r[1]),'totalGP':n(r[2]),'musterGP':n(r[3]),'dysfunctionalGP':n(r[4]),'labourAll':n(r[6]),'mrAll':n(r[7]),'ongoingAll':n(r[8]),'labourIndividual':n(r[10]),'mrIndividual':n(r[11]),'labourCommunity':n(r[12]),'mrCommunity':n(r[13]),'pmayOngoing':n(r[15]),'pmayMR':n(r[16]),'ekLabour':n(r[18]),'ekOngoing':n(r[19]),'ekMR':n(r[20])})

    # Sheet1 lower embedded table = Screen 2 authority.
    # Columns: GP, GP with progress, labour, works with MR, no-eKYC, Muster Rolls.
    header_idx=None
    for i,r in enumerate(svals):
        if len(r)>2 and 'total no. of gram panchayats' in c(r[2]).lower():
            header_idx=i; break
    daily=[]
    if header_idx is not None:
        for r in svals[header_idx+2:]:
            if len(r)<8: continue
            jan=janpad(r[1])
            if jan not in VALID_JANPADS: continue
            daily.append({'janpad':jan,'totalGP':n(r[2]),'gpsProgress':n(r[3]),'labour':n(r[4]),'worksMR':n(r[6]),'noEkyc':n(r[7]),'mrs':n(r[8])})
    if len(daily)!=8:
        raise SystemExit(f'Screen 2 table parse failed: expected 8 Janpads, got {len(daily)}')

    # V46: Screen-2 is authoritative for shared KPIs even when a full workbook is available.
    dmap={r['janpad']:r for r in daily}
    for o in official:
        d=dmap.get(o['janpad'])
        if not d: continue
        o['totalGP']=d['totalGP']; o['musterGP']=d['gpsProgress']; o['dysfunctionalGP']=max(0,d['totalGP']-d['gpsProgress'])
        o['labourAll']=d['labour']; o['mrAll']=d['worksMR']; o['noEkyc']=d['noEkyc']; o['mrs']=d['mrs']

    # VBG = work-level add-ons for engineer cards, category-wise work and expenditure buckets.
    v=wb['VBG']; wm={}; cm={}
    inst=re.compile(r'(school|prathmik|madhyamik|shala|vidyalaya|प्राथमिक|माध्यमिक|शाला|विद्यालय)',re.I)
    def exp_bucket(p):
        if p<=0: return 'b0'
        if p<=25: return 'b25'
        if p<=60: return 'b60'
        if p<=75: return 'b75'
        if p<=90: return 'b90'
        return 'b90p'
    for i,r in enumerate(v.iter_rows(values_only=True),start=1):
        if i<5 or len(r)<22: continue
        jan,gp,fy,status,code,name,wt=janpad(r[2]),up(r[3]),c(r[4]),c(r[5]).lower(),c(r[6]),c(r[7]),c(r[8])
        if not jan or not gp or not code or 'ongoing' not in status: continue
        eng,cl=gpmap.get((jan,gp),('Unmapped','Unmapped')); key=(jan,eng,cl,gp)
        z=wm.setdefault(key,{'janpad':jan,'engineer':eng,'cluster':cl,'panchayat':gp,'workTotal':0,'pmayOngoing':0,'ekOngoing':0,'currentFYActive':0})
        z['workTotal']+=1
        if 'pmay' in wt.lower(): z['pmayOngoing']+=1
        if final_category(name,wt,fy)=='Ek Bagiya': z['ekOngoing']+=1
        if n(r[21])>0: z['currentFYActive']+=1

        category=final_category(name,wt,fy)
        sanctioned=n(r[11]) if len(r)>11 else (n(r[9])+n(r[10]))
        booked=(n(r[16])+n(r[17])) if len(r)>17 else 0
        ep=(booked*100/sanctioned) if sanctioned>0 else 0
        ck=(jan,eng,cl,gp,category)
        q=cm.setdefault(ck,{'janpad':jan,'engineer':eng,'cluster':cl,'panchayat':gp,'category':category,'workCount':0,'totalSanction':0,'totalBooked':0,'b0':0,'b25':0,'b60':0,'b75':0,'b90':0,'b90p':0})
        q['workCount']+=1; q['totalSanction']+=sanctioned; q['totalBooked']+=booked; q[exp_bucket(ep)]+=1

    return {'title':title,'rows':rows,'official':official,'daily':daily,'workmix':list(wm.values()),'categorymix':list(cm.values()),'_sourceDates':{'RepDay':rep_date,'Sheet1':sheet1_date}}

path,source=obtain(); data=parse(path)
dates=data.pop('_sourceDates',{})
# V51 freshness lock: preserve updatedAt when the parsed workbook content is identical.
old_data=None
if OUT.exists():
    try:
        old_s=OUT.read_text(encoding='utf-8').strip()
        old_s=re.sub(r'^window\.AUTO_REPORT\s*=\s*','',old_s).rstrip(';')
        old_data=json.loads(old_s)
    except Exception:
        old_data=None
payload_keys=['title','rows','official','daily','workmix','categorymix']
old_payload={k:(old_data or {}).get(k) for k in payload_keys}
new_payload={k:data.get(k) for k in payload_keys}
changed=json.dumps(old_payload,ensure_ascii=False,sort_keys=True,separators=(',',':'))!=json.dumps(new_payload,ensure_ascii=False,sort_keys=True,separators=(',',':'))
old_meta=(old_data or {}).get('meta',{}) if isinstance(old_data,dict) else {}
updated_at=datetime.now(timezone.utc).isoformat() if changed or not old_meta.get('updatedAt') else old_meta.get('updatedAt')
data['meta']={'mode':'auto','status':'ok','updatedAt':updated_at,'source':source,'rowCount':len(data['rows']),'sourceDates':dates,'dataChangedOnLastFetch':changed}
OUT.write_text('window.AUTO_REPORT='+json.dumps(data,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')
print(f'Wrote {OUT} with {len(data["rows"])} GP rows and {len(data["daily"])} Screen-2 Janpad rows')
